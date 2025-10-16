import pandas as pd
import numpy as np
import json
from typing import Dict, List, Any, Optional, Tuple
import os
import tempfile
import shutil
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import io
import base64
import time
from openpyxl import load_workbook

# Stockage temporaire en mémoire
uploaded_files = {}
_UNIQUES_TTL_SECONDS = 15 * 60  # 15 minutes

# Helpers pour lecture légère et sélection d'engines
def _select_engine(path: str) -> Optional[str]:
    lower = path.lower()
    if lower.endswith('.xlsx'):
        return 'openpyxl'
    if lower.endswith('.xls'):
        return 'xlrd'
    return None

def _read_excel(path: str, nrows: Optional[int] = None, usecols: Optional[List[str]] = None) -> pd.DataFrame:
    engine = _select_engine(path)
    return pd.read_excel(path, nrows=nrows, usecols=usecols, engine=engine)

async def preview_excel(file):
    if not file.filename.endswith((".xls", ".xlsx")):
        return {"error": "Le fichier doit être un Excel (.xls ou .xlsx)"}
    
    # Sauvegarder le fichier en temporaire pour éviter de charger tout le DataFrame
    suffix = ".xlsx" if file.filename.lower().endswith(".xlsx") else ".xls"
    tmp_dir = tempfile.gettempdir()
    tmp_path = os.path.join(tmp_dir, f"preview_{next(tempfile._get_candidate_names())}{suffix}")

    file.file.seek(0)
    with open(tmp_path, "wb") as out:
        shutil.copyfileobj(file.file, out)

    # Lecture légère: colonnes + premières lignes
    sample_df = _read_excel(tmp_path, nrows=5)
    sample_df = sample_df.replace([np.nan, np.inf, -np.inf], None)

    uploaded_files[file.filename] = {
        "path": tmp_path,
        "df": None,  # chargé plus tard si nécessaire
        "columns": sample_df.columns.tolist(),
        "uniques_cache": {}
    }

    return {
        "filename": file.filename,
        "rows": int(len(sample_df)),
        "columns": sample_df.columns.tolist(),
        "preview": sample_df.head(5).to_dict(orient="records")
    }

async def select_columns(filename: str, variables_explicatives: List[str], variable_a_expliquer: List[str], selected_data: Dict = None):
    if filename not in uploaded_files:
        return {"error": "Fichier non trouvé. Faites d'abord /excel/preview."}
    
    file_ref = uploaded_files[filename]
    # Support rétro-compatible si ancien format (DataFrame direct)
    df = file_ref if isinstance(file_ref, pd.DataFrame) else file_ref.get("df")

    # Colonnes demandées
    all_columns = variables_explicatives + variable_a_expliquer

    # Identifier les colonnes restantes (celles qui ne sont ni explicatives ni à expliquer)
    # Déterminer les colonnes disponibles sans charger tout le DataFrame
    if isinstance(file_ref, dict) and "columns" in file_ref:
        all_df_columns = set([str(c) for c in file_ref["columns"]])
    else:
        all_df_columns = set(df.columns)
    remaining_columns = list(all_df_columns - set(all_columns))

    # Vérifier que les colonnes demandées existent sans nécessiter le DataFrame complet
    for col in all_columns:
        if col not in all_df_columns:
            return {"error": f"La colonne '{col}' n'existe pas dans {filename}"}

    # Si selected_data n'est pas fourni, retourner uniquement les noms des colonnes restantes (lazy load des valeurs)
    if selected_data is None:
        return {
            "filename": str(filename),
            "variables_explicatives": [str(col) for col in variables_explicatives],
            "variables_a_expliquer": [str(var) for var in variable_a_expliquer],
            "remaining_columns": [str(col) for col in remaining_columns],
            "remaining_data": {},
            "message": "Veuillez sélectionner les données des colonnes restantes sur lesquelles vous voulez travailler"
        }
    
    # Si selected_data est fourni, traiter la sélection finale (charger DF si nécessaire)
    if df is None and isinstance(file_ref, dict):
        try:
            df = _read_excel(file_ref["path"])
            file_ref["df"] = df
        except Exception as e:
            return {"error": f"Impossible de charger le fichier complet: {str(e)}"}

    # Re-sécurité si pour une raison quelconque df reste None
    if df is None:
        return {"error": "Données introuvables pour ce fichier. Veuillez relancer l'aperçu."}
    # Préparer les données explicatives
    X = df[variables_explicatives]
    
    # Préparer les variables à expliquer (chacune séparément)
    y_variables = {}
    for var in variable_a_expliquer:
        y_variables[var] = df[var]

    # Préparer les résultats pour chaque variable à expliquer
    results = []
    for var in variable_a_expliquer:
        # Convertir les données pandas en types Python natifs
        y_data = df[var]
        
        # Calculer les statistiques avec conversion en types natifs
        y_stats = {
            "count": int(y_data.count()),  # Convertir en int natif
            "mean": None,
            "std": None,
            "min": None,
            "max": None
        }
        
        # Vérifier si la colonne est numérique pour calculer les stats
        if y_data.dtype in ['int64', 'float64']:
            try:
                y_stats["mean"] = float(y_data.mean()) if not pd.isna(y_data.mean()) else None
                y_stats["std"] = float(y_data.std()) if not pd.isna(y_data.std()) else None
                y_stats["min"] = float(y_data.min()) if not pd.isna(y_data.min()) else None
                y_stats["max"] = float(y_data.max()) if not pd.isna(y_data.max()) else None
            except:
                # En cas d'erreur, garder None
                pass
        
        # Convertir les aperçus en types natifs
        y_preview = []
        for val in y_data.head(5):
            if pd.isna(val):
                y_preview.append(None)
            elif isinstance(val, (np.integer, np.floating)):
                y_preview.append(float(val) if isinstance(val, np.floating) else int(val))
            else:
                y_preview.append(str(val))
        
        result = {
            "variable_a_expliquer": str(var),  # Convertir en string natif
            "variables_explicatives": [str(col) for col in variables_explicatives],  # Convertir en strings natifs
            "X_preview": X.head(5).to_dict(orient="records"),
            "y_preview": y_preview,
            "y_stats": y_stats
        }
        results.append(result)

    # Préparer les données sélectionnées par l'utilisateur
    selected_data_with_columns = {}
    for col_name, selected_values in selected_data.items():
        if col_name in df.columns:
            # Filtrer le DataFrame pour ne garder que les lignes où la colonne contient les valeurs sélectionnées
            mask = df[col_name].isin(selected_values)
            filtered_df = df[mask]
            
            # Récupérer les données de cette colonne filtrée
            col_data = filtered_df[col_name].tolist()
            # Convertir en types Python natifs
            converted_col_data = []
            for val in col_data:
                if pd.isna(val):
                    converted_col_data.append(None)
                elif isinstance(val, (np.integer, np.floating)):
                    converted_col_data.append(float(val) if isinstance(val, np.floating) else int(val))
                else:
                    converted_col_data.append(str(val))
            
            selected_data_with_columns[str(col_name)] = converted_col_data

    return {
        "filename": str(filename),  # Convertir en string natif
        "variables_explicatives": [str(col) for col in variables_explicatives],  # Convertir en strings natifs
        "variables_a_expliquer": [str(var) for var in variable_a_expliquer],  # Convertir en strings natifs
        "selected_data": selected_data_with_columns,  # Données choisies par l'utilisateur avec noms de colonnes
        "results": results,
        "summary": {
            "total_variables_explicatives": int(len(variables_explicatives)),  # Convertir en int natif
            "total_variables_a_expliquer": int(len(variable_a_expliquer)),  # Convertir en int natif
            "total_rows": int(len(df)),  # Convertir en int natif
            "total_selected_columns": int(len(selected_data))  # Nombre de colonnes avec données sélectionnées
        }
    }

async def get_column_unique_values(filename: str, column_name: str, search: Optional[str] = None, offset: int = 0, limit: int = 200):
    if filename not in uploaded_files:
        return {"error": "Fichier non trouvé. Faites d'abord /excel/preview."}
    
    file_ref = uploaded_files[filename]
    # Cache des uniques par colonne
    if isinstance(file_ref, dict):
        cache = file_ref.setdefault("uniques_cache", {})
        cached = cache.get(column_name)
        if cached and (pd.Timestamp.now().timestamp() - cached.get("ts", 0) <= _UNIQUES_TTL_SECONDS):
            return {
                "filename": str(filename),
                "column_name": str(column_name),
                "unique_values": cached["values"],
                "total_unique_values": len(cached["values"])
            }
    # Si nous avons un DataFrame entier (ancien comportement)
    if isinstance(file_ref, pd.DataFrame):
        df = file_ref
        if column_name not in df.columns:
            return {"error": f"La colonne '{column_name}' n'existe pas dans {filename}"}
        series = df[column_name]
    else:
        # Lecture ciblée uniquement de la colonne depuis le disque
        path = file_ref.get("path")
        try:
            engine = _select_engine(path)
            if engine == 'openpyxl' and path.lower().endswith('.xlsx'):
                # Ultra-light streaming lecture avec openpyxl en mode read_only
                wb = load_workbook(path, read_only=True, data_only=True)
                ws = wb[wb.sheetnames[0]]

                # Trouver l'index de la colonne via l'entête déjà connue
                header_columns = file_ref.get("columns", [])
                try:
                    col_idx_zero_based = header_columns.index(column_name)
                except ValueError:
                    return {"error": f"La colonne '{column_name}' n'existe pas dans {filename}"}
                col_idx = col_idx_zero_based + 1  # openpyxl est 1-based

                # Itérer sur les valeurs de la colonne (à partir de la ligne 2 pour ignorer l'entête)
                seen_overall = set()
                collected = []
                filtered_seen = set()
                need = (offset or 0) + (limit or 200)
                search_lower = (search or "").lower()

                # Flag pour savoir si on a plus de résultats
                has_more_flag = False

                start_time = time.perf_counter()
                max_seconds = 6.0  # budget temps pour éviter timeouts plateformes gratuites
                skipped = 0
                for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx, values_only=True):
                    # Stop si on a un budget temps dépassé mais déjà des résultats
                    if (time.perf_counter() - start_time) > max_seconds and len(collected) > 0:
                        has_more_flag = True
                        break
                    cell_value = row[0]
                    # Normalisation similaire au chemin pandas
                    if cell_value is None:
                        normalized = None
                        display = ""
                    elif isinstance(cell_value, (int, float, np.integer, np.floating)):
                        # Convertir floats numpy vs natifs
                        if isinstance(cell_value, float) and cell_value.is_integer():
                            normalized = int(cell_value)
                        else:
                            normalized = float(cell_value) if isinstance(cell_value, (float, np.floating)) else int(cell_value)
                        display = str(normalized)
                    elif isinstance(cell_value, bool):
                        normalized = str(cell_value)
                        display = normalized
                    else:
                        normalized = str(cell_value)
                        display = normalized

                    # Unicité globale
                    key = display if normalized is not None else ""
                    if key in seen_overall:
                        continue
                    seen_overall.add(key)

                    # Filtre de recherche
                    if search_lower:
                        if display.lower().find(search_lower) == -1:
                            continue

                    # Unicité dans le set filtré
                    if key in filtered_seen:
                        continue
                    filtered_seen.add(key)

                    # Gérer offset
                    if skipped < (offset or 0):
                        skipped += 1
                        continue

                    # Ajouter jusqu'à limit
                    collected.append(normalized)
                    if len(collected) >= (limit or 200):
                        # On a ce qu'il faut; vérifier s'il reste potentiellement plus
                        has_more_flag = True
                        break

                wb.close()

                # Retourner rapidement sans construire toute la liste d'unicité
                return {
                    "filename": str(filename),
                    "column_name": str(column_name),
                    "unique_values": collected,
                    "total_unique_values": -1,  # inconnu (non utilisé côté front)
                    "filtered_total_unique_values": (offset or 0) + len(collected) + (1 if has_more_flag else 0),
                    "offset": int(offset or 0),
                    "limit": int(limit or 200),
                    "has_more": has_more_flag
                }
            else:
                # Fallback: pandas lecture d'une seule colonne
                col_df = pd.read_excel(path, usecols=[column_name], engine=engine)
                if column_name not in col_df.columns:
                    return {"error": f"La colonne '{column_name}' n'existe pas dans {filename}"}
                series = col_df[column_name]
        except Exception as e:
            return {"error": f"Lecture colonne échouée: {str(e)}"}

    # Récupérer toutes les valeurs uniques de la colonne (base)
    base_unique_values = series.dropna().unique()

    # Convertir en types Python natifs (base)
    base_converted_values = []
    for val in base_unique_values:
        if pd.isna(val):
            base_converted_values.append(None)
        elif isinstance(val, (np.integer, np.floating)):
            base_converted_values.append(float(val) if isinstance(val, np.floating) else int(val))
        else:
            base_converted_values.append(str(val))

    # Filtrage (search)
    if search:
        search_lower = search.lower()
        filtered_values = [v for v in base_converted_values if ("" if v is None else str(v)).lower().find(search_lower) != -1]
    else:
        filtered_values = base_converted_values

    total_filtered = len(filtered_values)
    start = max(0, int(offset))
    end = max(start, start + int(limit)) if limit is not None else total_filtered
    paged_values = filtered_values[start:end]

    result = {
        "filename": str(filename),
        "column_name": str(column_name),
        "unique_values": paged_values,
        "total_unique_values": len(base_converted_values),
        "filtered_total_unique_values": total_filtered,
        "offset": start,
        "limit": int(limit),
        "has_more": end < total_filtered
    }

    # Mettre en cache la base complète (non filtrée)
    if isinstance(file_ref, dict):
        cache[column_name] = {"values": base_converted_values, "ts": pd.Timestamp.now().timestamp()}

    return result

# ============================================================================
# NOUVELLES FONCTIONS POUR L'ARBRE DE DÉCISION
# ============================================================================

def calculate_percentage_variance(df: pd.DataFrame, explanatory_var: str, target_var: str, target_value: Any) -> float:
    """
    Calcule l'écart-type des pourcentages des valeurs d'une variable explicative
    pour une valeur cible donnée.
    
    CORRECTION: Les pourcentages sont calculés par rapport au total des accidents
    de chaque valeur de la variable explicative, pas par rapport au total filtré.
    """
    try:
        # Filtrer pour la valeur cible (exclure les NaN)
        target_mask = (df[target_var] == target_value) & (df[target_var].notna())
        filtered_df = df[target_mask]
        
        if len(filtered_df) == 0:
            return 0.0
        
        # Obtenir toutes les valeurs uniques de la variable explicative dans le dataset filtré
        all_explanatory_values = df[explanatory_var].dropna().unique()
        
        if len(all_explanatory_values) == 0:
            return 0.0
        
        # Pour chaque valeur de la variable explicative, calculer le pourcentage
        # d'accidents de type "target_value" parmi tous les accidents de cette valeur
        percentages = []
        
        for explanatory_value in all_explanatory_values:
            # Nombre total d'accidents avec cette valeur explicative
            total_explanatory = len(df[df[explanatory_var] == explanatory_value])
            
            # Nombre d'accidents avec cette valeur explicative ET la valeur cible
            target_and_explanatory = len(
                df[(df[explanatory_var] == explanatory_value) & 
                   (df[target_var] == target_value) & 
                   (df[target_var].notna())]
            )
            
            # Calculer le pourcentage
            if total_explanatory > 0:
                percentage = (target_and_explanatory / total_explanatory) * 100
                percentages.append(percentage)
        
        # Calculer l'écart-type des pourcentages
        if len(percentages) > 1:
            return float(np.std(percentages))
        else:
            return 0.0
            
    except Exception as e:
        return 0.0

def select_best_explanatory_variable(df: pd.DataFrame, available_vars: List[str], 
                                   target_var: str, target_value: Any) -> Tuple[str, float]:
    """
    Sélectionne la variable explicative avec le plus grand écart-type des pourcentages.
    """
    best_var = None
    best_variance = -1
    
    var_variances = {}
    for var in available_vars:
        variance = calculate_percentage_variance(df, var, target_var, target_value)
        var_variances[var] = variance
    
    # Sélectionner la variable avec la plus grande variance
    if var_variances:
        best_var = max(var_variances, key=var_variances.get)
        best_variance = var_variances[best_var]
    
    return best_var, best_variance

def calculate_branch_percentages(df: pd.DataFrame, explanatory_var: str, 
                               target_var: str, target_value: Any) -> Dict[str, Dict[str, Any]]:
    """
    Calcule les pourcentages et comptages pour chaque branche d'une variable explicative.
    
    CORRECTION: Les pourcentages sont calculés par rapport au total des accidents
    de chaque valeur de la variable explicative, pas par rapport au total filtré.
    """
    try:

        
        # Obtenir toutes les valeurs uniques de la variable explicative dans le dataset filtré
        all_explanatory_values = df[explanatory_var].dropna().unique()
        
        if len(all_explanatory_values) == 0:
            return {}
        
        branches = {}
        
        for explanatory_value in all_explanatory_values:
            # Nombre total d'accidents avec cette valeur explicative
            total_explanatory = len(df[df[explanatory_var] == explanatory_value])
            
            # Nombre d'accidents avec cette valeur explicative ET la valeur cible
            target_and_explanatory = len(
                df[(df[explanatory_var] == explanatory_value) & 
                   (df[target_var] == target_value) & 
                   (df[target_var].notna())]
            )
            
            # Calculer le pourcentage
            if total_explanatory > 0:
                percentage = (target_and_explanatory / total_explanatory) * 100
                branches[str(explanatory_value)] = {
                    "count": int(target_and_explanatory),
                    "percentage": round(percentage, 2),
                    "subtree": None  # Sera rempli récursivement
                }
        
        return branches
        
    except Exception as e:
        return {}

def construct_tree_for_value(df: pd.DataFrame, target_value: Any, target_var: str, 
                           available_explanatory_vars: List[str], current_path: List[str] = None) -> Dict[str, Any]:
    """
    Construit récursivement l'arbre de décision pour une valeur cible donnée.
    """
    if current_path is None:
        current_path = []
    
    # Critère d'arrêt : plus de variables explicatives disponibles
    if not available_explanatory_vars:
        return {
            "type": "leaf",
            "message": "Plus de variables explicatives disponibles"
        }
    
    # Sélectionner la meilleure variable explicative
    best_var, best_variance = select_best_explanatory_variable(
        df, available_explanatory_vars, target_var, target_value
    )
    
    if best_var is None:
        return {
            "type": "leaf",
            "message": "Aucune variable explicative valide trouvée"
        }
    
    # Calculer les branches pour cette variable
    branches = calculate_branch_percentages(df, best_var, target_var, target_value)
    
    # Créer le nœud de l'arbre
    tree_node = {
        "type": "node",
        "variable": best_var,
        "variance": round(best_variance, 4),
        "branches": branches,
        "path": current_path + [best_var]
    }
    
    # Variables explicatives restantes pour les sous-arbres
    remaining_vars = [var for var in available_explanatory_vars if var != best_var]
    
    # Construire récursivement les sous-arbres pour chaque branche
    for branch_value, branch_data in branches.items():
        # Filtrer le DataFrame pour cette branche
        # Convertir branch_value en type approprié pour la comparaison
        if branch_value == 'False':
            branch_value_converted = False
        elif branch_value == 'True':
            branch_value_converted = True
        else:
            branch_value_converted = branch_value
        
        branch_mask = (df[best_var] == branch_value_converted) & (df[best_var].notna())
        filtered_df = df[branch_mask]
        
        if len(filtered_df) > 0 and remaining_vars:
            # Construire le sous-arbre récursivement
            subtree = construct_tree_for_value(
                filtered_df, target_value, target_var, 
                remaining_vars, current_path + [best_var, branch_value]
            )
            branch_data["subtree"] = subtree
    
    return tree_node

async def build_decision_tree(filename: str, variables_explicatives: List[str], 
                            variables_a_expliquer: List[str], selected_data: Dict[str, Any]) -> Dict[str, Any]:
    """
    Construit l'arbre de décision complet pour toutes les variables à expliquer.
    """
    if filename not in uploaded_files:
        return {"error": "Fichier non trouvé. Faites d'abord /excel/preview."}
    
    file_ref = uploaded_files[filename]
    # Charger DataFrame à la demande si nous avons un chemin
    if isinstance(file_ref, dict):
        df = file_ref.get("df")
        if df is None:
            try:
                df = _read_excel(file_ref.get("path"))
                file_ref["df"] = df
            except Exception as e:
                return {"error": f"Impossible de charger le fichier complet: {str(e)}"}
    else:
        df = file_ref
    
    # Étape 1: Filtrer l'échantillon initial basé sur les variables restantes sélectionnées
    
    # Identifier les colonnes restantes (ni explicatives ni à expliquer)
    all_columns = variables_explicatives + variables_a_expliquer
    remaining_columns = [col for col in df.columns if col not in all_columns]
    
    # Filtrer pour les variables restantes sélectionnées
    initial_mask = pd.Series([True] * len(df), index=df.index)
    
    for col_name, selected_values in selected_data.items():
        if col_name in remaining_columns and selected_values:
            # Conversion automatique des types pour la correspondance
            converted_values = []
            for val in selected_values:
                if isinstance(val, str):
                    if val.lower() == 'true':
                        converted_values.append(True)
                    elif val.lower() == 'false':
                        converted_values.append(False)
                    else:
                        converted_values.append(val)
                else:
                    converted_values.append(val)
            
            col_mask = df[col_name].isin(converted_values)
            initial_mask = initial_mask & col_mask
    
    filtered_df = df[initial_mask].copy()
    
    # Analyser l'impact du filtrage sur les variables explicatives
    filtering_analysis = analyze_sample_filtering_impact(df, filtered_df, variables_explicatives)
    
    # Étape 2: Construire l'arbre pour chaque variable à expliquer
    
    decision_trees = {}
    
    for target_var in variables_a_expliquer:
        # IMPORTANT: Utiliser seulement les valeurs SÉLECTIONNÉES, pas toutes les valeurs uniques
        if target_var in selected_data and selected_data[target_var]:
            # Utiliser les valeurs sélectionnées par l'utilisateur
            target_values = selected_data[target_var]
        else:
            # Fallback: utiliser toutes les valeurs uniques si aucune sélection
            target_values = filtered_df[target_var].dropna().unique()
        
        target_trees = {}
        
        for target_value in target_values:
            # Construire l'arbre pour cette valeur
            tree = construct_tree_for_value(
                filtered_df, target_value, target_var, 
                variables_explicatives.copy(), []
            )
            
            target_trees[str(target_value)] = tree
        
        decision_trees[target_var] = target_trees
    
    return {
        "filename": filename,
        "variables_explicatives": variables_explicatives,
        "variables_a_expliquer": variables_a_expliquer,
        "filtered_sample_size": len(filtered_df),
        "original_sample_size": len(df),
        "decision_trees": decision_trees
    }

def generate_tree_pdf(decision_trees: Dict[str, Any], filename: str) -> str:
    """
    Génère un PDF de l'arbre de décision avec structure arborescente claire et branches gauche/droite.
    """
    try:
        # Créer un buffer en mémoire pour le PDF
        buffer = io.BytesIO()
        
        # Créer le document PDF
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            spaceAfter=25,
            alignment=TA_CENTER,
            textColor=colors.darkblue
        )
        
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Heading2'],
            fontSize=16,
            spaceAfter=20,
            textColor=colors.darkgreen
        )
        
        node_style = ParagraphStyle(
            'NodeStyle',
            parent=styles['Normal'],
            fontSize=12,
            spaceAfter=8,
            textColor=colors.darkblue,
            leftIndent=20
        )
        
        branch_style = ParagraphStyle(
            'BranchStyle',
            parent=styles['Normal'],
            fontSize=11,
            spaceAfter=6,
            textColor=colors.purple,
            leftIndent=40
        )
        
        leaf_style = ParagraphStyle(
            'LeafStyle',
            parent=styles['Normal'],
            fontSize=10,
            spaceAfter=4,
            textColor=colors.darkgreen,
            leftIndent=60
        )
        
        # Titre principal
        story.append(Paragraph("🌳 ARBRE DE DÉCISION - ANALYSE STATISTIQUE", title_style))
        story.append(Spacer(1, 25))
        
        # Informations du fichier
        story.append(Paragraph(f"📁 <b>Fichier:</b> {filename}", styles['Normal']))
        story.append(Spacer(1, 15))
        
        # Fonction récursive pour afficher l'arbre avec structure claire
        def add_tree_to_story(node, level=0, path=""):
            try:
                if node.get("type") == "leaf":
                    # Feuille de l'arbre
                    story.append(Paragraph(f"🍃 {node.get('message', 'Fin de branche')}", leaf_style))
                else:
                    # Nœud principal avec variable explicative
                    indent = "&nbsp;" * (level * 8)
                    story.append(Paragraph(
                        f"{indent}🌿 <b>{node['variable']}</b> (Écart-type: {node['variance']})", 
                        node_style
                    ))
                    
                    # Branches avec structure gauche/droite
                    branches = list(node['branches'].items())
                    mid_point = len(branches) // 2
                    
                    # Branches gauches
                    if mid_point > 0:
                        story.append(Paragraph(f"{indent}&nbsp;&nbsp;├─ <b>BRANCHES GAUCHES:</b>", branch_style))
                        for i, (branch_value, branch_data) in enumerate(branches[:mid_point]):
                            story.append(Paragraph(
                                f"{indent}&nbsp;&nbsp;&nbsp;&nbsp;├─ <b>{branch_value}</b>: {branch_data['count']} ({branch_data['percentage']}%)", 
                                branch_style
                            ))
                            
                            # Sous-arbre récursif
                            if branch_data.get('subtree'):
                                add_tree_to_story(branch_data['subtree'], level + 1, f"{path} → {branch_value}")
                    
                    # Branches droites
                    if len(branches) > mid_point:
                        story.append(Paragraph(f"{indent}&nbsp;&nbsp;├─ <b>BRANCHES DROITES:</b>", branch_style))
                        for i, (branch_value, branch_data) in enumerate(branches[mid_point:]):
                            story.append(Paragraph(
                                f"{indent}&nbsp;&nbsp;&nbsp;&nbsp;├─ <b>{branch_value}</b>: {branch_data['count']} ({branch_data['percentage']}%)", 
                                branch_style
                            ))
                            
                            # Sous-arbre récursif
                            if branch_data.get('subtree'):
                                add_tree_to_story(branch_data['subtree'], level + 1, f"{path} → {branch_value}")
                    
                    story.append(Spacer(1, 10))
            except Exception as e:
                story.append(Paragraph(f"❌ Erreur lors de l'affichage du nœud", styles['Normal']))
        
        # Pour chaque variable à expliquer
        for target_var, target_trees in decision_trees.items():
            try:
                story.append(Paragraph(f"🎯 <b>VARIABLE À EXPLIQUER: {target_var}</b>", subtitle_style))
                story.append(Spacer(1, 15))
                
                # Pour chaque valeur de cette variable
                for target_value, tree in target_trees.items():
                    try:
                        story.append(Paragraph(f"📊 <b>VALEUR CIBLE: {target_value}</b>", styles['Heading3']))
                        story.append(Spacer(1, 10))
                        
                        # Construire l'arbre récursivement
                        add_tree_to_story(tree, 0, target_value)
                        story.append(Spacer(1, 20))
                    except Exception as e:
                        story.append(Paragraph(f"❌ Erreur lors du traitement de la valeur {target_value}", styles['Normal']))
                
                story.append(Spacer(1, 25))
            except Exception as e:
                story.append(Paragraph(f"❌ Erreur lors du traitement de la variable {target_var}", styles['Normal']))
        
        # Construire le PDF
        doc.build(story)
        
        # Obtenir le contenu du buffer
        pdf_content = buffer.getvalue()
        buffer.close()
        
        # Encoder en base64
        pdf_base64 = base64.b64encode(pdf_content).decode('utf-8')
        
        return pdf_base64
        
    except Exception as e:
        return ""

async def build_decision_tree_with_pdf(filename: str, variables_explicatives: List[str], 
                                     variables_a_expliquer: List[str], selected_data: Dict[str, Any]) -> Dict[str, Any]:
    """
    Construit l'arbre de décision et génère le PDF correspondant.
    """
    # Construire l'arbre
    tree_result = await build_decision_tree(filename, variables_explicatives, variables_a_expliquer, selected_data)
    
    if "error" in tree_result:
        return tree_result
    
    # Générer le PDF
    pdf_base64 = generate_tree_pdf(tree_result["decision_trees"], filename)
    
    if pdf_base64:
        tree_result["pdf_base64"] = pdf_base64
        tree_result["pdf_generated"] = True
    else:
        tree_result["pdf_generated"] = False
    
    return tree_result

def analyze_sample_filtering_impact(df: pd.DataFrame, filtered_df: pd.DataFrame, 
                                   variables_explicatives: List[str]) -> Dict[str, Any]:
    """
    Analyse l'impact du filtrage de l'échantillon sur les variables explicatives.
    Retourne des avertissements et suggestions pour l'utilisateur.
    """
    warnings = []
    suggestions = []
    
    for var in variables_explicatives:
        original_unique = df[var].nunique()
        filtered_unique = filtered_df[var].nunique()
        
        if filtered_unique == 1:
            warnings.append(f"⚠️ Variable '{var}' n'a plus qu'une seule valeur unique dans l'échantillon filtré")
            suggestions.append(f"Considérez élargir la sélection pour '{var}' ou la retirer des variables explicatives")
        elif filtered_unique < original_unique * 0.5:
            warnings.append(f"⚠️ Variable '{var}' a perdu plus de 50% de ses valeurs uniques")
            suggestions.append(f"La variable '{var}' pourrait avoir une variance réduite")
        elif filtered_unique < 3:
            warnings.append(f"⚠️ Variable '{var}' a moins de 3 valeurs uniques dans l'échantillon filtré")
            suggestions.append(f"La variable '{var}' pourrait avoir une variance faible")
    
    return {
        "warnings": warnings,
        "suggestions": suggestions,
        "original_sample_size": len(df),
        "filtered_sample_size": len(filtered_df),
        "reduction_percentage": round(((len(df) - len(filtered_df)) / len(df) * 100), 1)
    }
